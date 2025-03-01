# 自动化获取boss招聘网的信息
# 引入By Class，辅助元素定位

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
import pandas as pd
import undetected_chromedriver as uc
import random
import re
import openpyxl

# 待查公司的Excel文件地址
company_name_file_path = 'E:/boss_zhipin/medical_company_name.xlsx'
# 爬取结果保存excel文件地址
save_file_path = 'E:/boss_zhipin/medical_company_info.xlsx'
# 保存到的sheet名称
save_sheet_name = 'company'

# 加载待查公司的Excel文件
wb = openpyxl.load_workbook(company_name_file_path)

# 选择默认的工作表（通常是第一个工作表）
sheet = wb.active

# 读取第一列的值
column_values = [cell.value for cell in sheet['A']]

# BOSS招聘网站
url = "https://www.zhipin.com/?city=100010000&ka=city-sites-100010000"
# 创建谷歌浏览器
browser = uc.Chrome()
# 打开网页
browser.get(url=url)
browser.set_window_size(1200, 800)  # 改变页面大小
# 等待10秒钟，最好别访问太快
sleep(10)

# 定义一个空数组，后面用来存储公司数据
company_lis = []

for search_company_name in column_values:
    # 选择搜索框
    searching = browser.find_element(By.XPATH,
                                     '//input[@type="text" and @name="query" and @class="ipt-search" and @placeholder="搜索职位、公司"]')
    # 输入，搜素
    searching.send_keys(search_company_name)
    # 点击搜索，通过回车点击
    searching.send_keys(Keys.ENTER)
    # 等待10秒
    sleep(10)

    # 页面跳转后定位窗口
    browser.switch_to.window(browser.window_handles[-1])
    # 查找公司简介的链接
    page_link = browser.find_elements(By.XPATH,
                                      '//div[@class="job-card-body clearfix"]/div[@class="job-card-right"]/div[@class="company-info"]/h3[@class="company-name"]/a[@ka="search_list_company_1_custompage"]')
    link = page_link[0]
    browser.get(url=link.get_attribute('href'))
    sleep(5)

    # 定位到包含所需信息的元素
    info_primary = browser.find_element(By.XPATH,
                                        "//div[@class='inner home-inner']/div/div[@class='info-primary']")
    # 提取公司名字
    company_name = info_primary.find_element(By.XPATH, ".//div[@class='info']/h1[@class='name']").text
    company_name = company_name.split('收藏')[0].strip()  # 移除公司名后面跟随的“收藏”

    info_text = info_primary.find_element(By.XPATH, ".//div[@class='info']/p").text
    match = re.search(r'^(\D+)(\d+-\d+人)(\D+)$', info_text)
    if match:
        financing_info = match.group(1)  # 提取公司资金情况
        # print('公司资金情况', financing_info)
        staff_count = match.group(2)  # 提取公司人数
        # print('公司人数', staff_count)
        industry = match.group(3)  # 提取industry
        # print('公司产业', industry)
        info = [company_name, financing_info, staff_count, industry]
    else:
        match = re.search(r'^(\D+)(\d+人以上)(\D+)$', info_text)
        if match:
            financing_info = match.group(1)  # 提取公司资金情况
            # print('公司资金情况', financing_info)
            staff_count = match.group(2)  # 提取公司人数
            # print('公司人数', staff_count)
            industry = match.group(3)  # 提取industry
            # print('公司产业', industry)
            info = [company_name, financing_info, staff_count, industry]
        else:
            info = [company_name, info_text, 'NAN', 'NAN']

    # 定位到包含所需信息的元素
    job_box = browser.find_element(By.CLASS_NAME, "job-box")

    # 提取公司简介
    try:
        company_intro = job_box.find_element(By.XPATH,
                                             ".//div[@class='job-sec'][h3='公司简介']/div[@class='text fold-text']").text
    except:
        company_intro = 'NAN'

    # 提取成立时间
    try:
        established_time = job_box.find_element(By.XPATH, ".//li[@class='business-detail-time']").text
        established_time = established_time.split('\n')[-1]
    except:
        established_time = 'NAN'

    # 提取注册资本
    try:
        registered_capital = job_box.find_element(By.XPATH, ".//li[@class='business-detail-money']").text
        registered_capital = registered_capital.split('\n')[-1]
    except:
        registered_capital = 'NAN'

    # 提取公司地址
    # company_address = job_box.find_element(By.XPATH, ".//div[@class='job-location']/div[@class='location-item show-map']/div[@class='location-address']").text
    company_location = job_box.find_element(By.XPATH, ".//div[@class='job-location']")
    address_list = company_location.find_elements(By.XPATH, ".//div[@class='location-address']")
    company_address = ""
    for address in address_list:
        company_address = company_address + address.text if company_address == "" else company_address + '\n' + address.text

    # 输出信息
    # print("公司简介：", company_intro)
    # print("成立时间：", established_time)
    # print("注册资本：", registered_capital)
    # print("公司地址：", company_address)

    company_info = info + [company_intro, established_time, registered_capital, company_address]
    company_lis.append(company_info)

    # -------------保存爬取的数据到excel------------------------
    # 获取指定的文件
    wb = openpyxl.load_workbook(save_file_path)
    # 获取指定的sheet
    ws = wb[save_sheet_name]
    # 获得最大行数
    max_row_num = ws.max_row
    # 获得最大列数
    max_col_num = ws.max_column

    # 将当前行设置为最大行数
    ws._current_row = max_row_num

    ws.append(company_info)

    # 保存文件
    wb.save(save_file_path)

    print('finish 公司名字：', company_name)

browser.close()
browser.quit()

print('finish all')
